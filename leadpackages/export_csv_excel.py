"""
CSV- + Excel-Export für Datenpakete (Entscheidung: CSV + Excel).
Gleiche CSV-Konventionen wie db_evaluated.export_csv() (Semikolon + UTF-8-BOM
für deutsches Excel), zusätzlich ein natives .xlsx über openpyxl.
"""
from __future__ import annotations

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font

_SPALTEN = [
    ("Firma", "name"),
    ("Branche", "branche"),
    ("Stadt", "stadt"),
    ("Region", "region"),
    ("Land", "land"),
    ("Adresse", "adresse"),
    ("Telefon", "telefon"),
    ("E-Mail", "email_adresse"),
    ("Website", "website_url"),
    ("Datenqualität (0-100)", "quality_score"),
    ("Potenzial", "potential_label"),
]


def art14_hinweis() -> str:
    """DSGVO Art. 14 / 26 — Pflicht-Hinweisblatt für den Käufer roher Firmendaten.
    Der Käufer wird zum Verantwortlichen und muss die Betroffenen über die Datenherkunft
    informieren (und darf sie nicht ohne Rechtsgrundlage für Werbung nutzen)."""
    return (
        "DATENSCHUTZ- UND NUTZUNGSHINWEIS (bitte vor Verwendung lesen)\n"
        "\n"
        "Diese Datei enthält geschäftliche Kontaktdaten, die aus öffentlich zugänglichen Quellen "
        "(Branchenverzeichnisse, Kartendienste, Firmen-Websites) zusammengestellt wurden.\n"
        "\n"
        "Mit Erhalt dieser Daten werden Sie datenschutzrechtlich Verantwortlicher (Art. 4 Nr. 7 "
        "DSGVO) für Ihre weitere Verarbeitung. Daraus folgt insbesondere:\n"
        "- Informationspflicht (Art. 14 DSGVO): Sie müssen die betroffenen Personen/Betriebe "
        "  spätestens bei der ersten Ansprache über Herkunft und Zweck der Verarbeitung informieren.\n"
        "- Werberechtliche Grenzen (UWG): Kalt-Telefon-/E-Mail-/Fax-Werbung ist ohne (mutmaßliche) "
        "  Einwilligung regelmäßig unzulässig. Prüfen Sie die Rechtsgrundlage vor jeder Nutzung.\n"
        "- Betroffenenrechte: Auskunft, Berichtigung, Löschung und Widerspruch sind zu beachten.\n"
        "\n"
        "Es wird keine Rechtsberatung erteilt; die rechtskonforme Nutzung liegt in Ihrer "
        "Verantwortung."
    )


def _val(row: dict, key: str) -> str:
    v = row.get(key, "")
    return "" if v is None else str(v)


def to_csv(rows: list[dict]) -> str:
    buf = io.StringIO()
    buf.write("﻿")
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    w.writerow([titel for titel, _ in _SPALTEN])
    for row in rows:
        w.writerow([_val(row, key) for _, key in _SPALTEN])
    return buf.getvalue()


def to_excel_bytes(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datenpaket"
    ws.append([titel for titel, _ in _SPALTEN])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_val(row, key) for _, key in _SPALTEN])
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 45)
    # Zweites Blatt: verpflichtender Datenschutz-/Nutzungshinweis (Art. 14 DSGVO).
    ws2 = wb.create_sheet("Datenschutzhinweis")
    for i, zeile in enumerate(art14_hinweis().split("\n"), start=1):
        c = ws2.cell(row=i, column=1, value=zeile)
        if i == 1:
            c.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 100
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
